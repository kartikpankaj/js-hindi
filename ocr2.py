import os
import json
import re
import openpyxl
from google.cloud import vision
from google.cloud import storage
from google.cloud.vision_v1 import types
from datetime import datetime

# Set the path to your service account JSON key
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "metal-repeater-436911-u7-233a3ddf9417.json"

# Replace with your bucket name and folder path
bucket_name = 'bucket_election'
folder_path = r'E:/Election/Test'  # Specify the folder path containing the PDF files


def delete_blob_if_exists(bucket_name, blob_name):
    """Deletes the blob if it already exists in the bucket."""
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    
    if blob.exists():
        blob.delete()
        print(f"Deleted existing blob {blob_name} from bucket {bucket_name}.")


def upload_to_bucket(bucket_name, file_path):
    """Uploads a file to the bucket, using the file's original name."""
    destination_blob_name = os.path.basename(file_path)
    delete_blob_if_exists(bucket_name, destination_blob_name)
    
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)
    blob.upload_from_filename(file_path)
    print(f"File {file_path} uploaded to {destination_blob_name} in bucket {bucket_name}.")


def perform_ocr(bucket_name, file_name):
    """Performs OCR on a PDF file stored in Google Cloud Storage."""
    client = vision.ImageAnnotatorClient()
    gcs_source_uri = f"gs://{bucket_name}/{file_name}"
    gcs_source = types.GcsSource(uri=gcs_source_uri)
    input_config = types.InputConfig(gcs_source=gcs_source, mime_type="application/pdf")

    gcs_destination_uri = f"gs://{bucket_name}/output/ocr/"
    gcs_destination = types.GcsDestination(uri=gcs_destination_uri)
    output_config = types.OutputConfig(gcs_destination=gcs_destination, batch_size=1)

    async_request = types.AsyncAnnotateFileRequest(
        features=[types.Feature(type=vision.Feature.Type.DOCUMENT_TEXT_DETECTION)],
        input_config=input_config,
        output_config=output_config,
    )

    operation = client.async_batch_annotate_files(requests=[async_request])
    print("Waiting for the OCR operation to complete...")
    operation.result(timeout=300)
    print("OCR processing completed.")


def download_and_save_results(bucket_name, output_text_file, file_name):
    """Downloads OCR results and saves the text to a file."""
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blobs = list(bucket.list_blobs(prefix="output/ocr/"))

    extracted_names = []

    with open(output_text_file, "a", encoding="utf-8") as file:
        for blob in blobs:
            json_data = blob.download_as_text()
            response = json.loads(json_data)

            for page in response['responses']:
                if 'fullTextAnnotation' in page:
                    full_text = page['fullTextAnnotation']['text']
                    file.write(f"\n--- Text from {file_name} ---\n")
                    file.write(full_text)
                    file.write("\n\n")
                    
                    # Use a revised regex pattern without look-behind
                    names = re.findall(r"नाव\s*:?\s*(\S+)", full_text)
                    extracted_names.extend(names)
                else:
                    file.write("No text detected on this page.\n")

    print(f"OCR text from {file_name} saved to {output_text_file}")
    return extracted_names


def save_names_to_excel(all_names, output_excel_file):
    """Saves extracted names from multiple PDFs to an Excel file, each PDF's names in a new sheet."""
    workbook = openpyxl.Workbook()
    for file_name, names in all_names.items():
        sheet = workbook.create_sheet(title=file_name[:31])  # Excel sheet names limited to 31 characters
        sheet["A1"] = "Names"

        for index, name in enumerate(names, start=2):
            sheet[f"A{index}"] = name

    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']  # Remove the default sheet created by openpyxl
    
    workbook.save(output_excel_file)
    print(f"Extracted names saved to {output_excel_file}")


# Iterate over each PDF in the specified folder
all_extracted_names = {}
current_date_time = datetime.now()
print("Started At:", current_date_time)
for file_name in os.listdir(folder_path):
    if file_name.endswith('.pdf'):
        file_path = os.path.join(folder_path, file_name)
        
        # Generate output file names based on the input file name
        output_text_file = f"output/ocr/{os.path.splitext(file_name)[0]}.txt"
        output_excel_file = f"output/excel/{os.path.splitext(file_name)[0]}.xlsx"
        
        # Ensure the directories exist before saving
        os.makedirs(os.path.dirname(output_text_file), exist_ok=True)
        os.makedirs(os.path.dirname(output_excel_file), exist_ok=True)

        # Upload the PDF to Google Cloud Storage
        upload_to_bucket(bucket_name, file_path)

        # Perform OCR on the uploaded PDF
        perform_ocr(bucket_name, file_name)

        # Download and save the OCR results, and extract names
        extracted_names = download_and_save_results(bucket_name, output_text_file, file_name)
        
        # Store extracted names for each file in a dictionary
        #all_extracted_names[file_name] = extracted_names

        # Save all extracted names to an Excel file
        #save_names_to_excel(all_extracted_names, output_excel_file)
current_date_time = datetime.now()
print("Completed At:", current_date_time)
